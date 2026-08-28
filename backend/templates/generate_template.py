from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def main() -> None:
    base = Path(__file__).parent
    base.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Harian
    ws1 = wb.active
    ws1.title = "Harian"
    headers1 = [
        "Tanggal",
        "Keterangan",
        "Kategori",
        "Supplier",
        "Item",
        "Qty",
        "Satuan",
        "Harga",
        "Masuk",
        "Keluar",
        "Saldo",
    ]
    ws1.append(headers1)

    # Sheet 2: Mingguan
    ws2 = wb.create_sheet("Mingguan")
    headers2 = [
        "Minggu",
        "Total Porsi",
        "Pendapatan",
        "Pengeluaran Bahan",
        "Upah",
        "Ops",
        "Laba",
    ]
    ws2.append(headers2)

    # Sheet 3: Bulanan
    ws3 = wb.create_sheet("Bulanan")
    headers3 = [
        "Bulan",
        "Pendapatan Bruto",
        "PPh22",
        "Pendapatan Bersih",
        "Total Bahan",
        "Total Upah",
        "Total Ops",
        "Laba Bersih",
        "Margin %",
    ]
    ws3.append(headers3)

    # Sheet 4: Stok
    ws4 = wb.create_sheet("Stok")
    headers4 = [
        "Nama Bahan",
        "Satuan",
        "Stok Awal",
        "Total Masuk",
        "Total Keluar",
        "Stok Akhir",
    ]
    ws4.append(headers4)

    # Sheet 5: Laporan Pemerintah (Breakdown: Bahan : Ops : Profit)
    ws5 = wb.create_sheet("Laporan Pemerintah")
    headers5 = [
        "Tanggal",
        "Bahan Baku (Rp)",
        "Operasional (Rp)",
        "Profit (Rp)",
        "Total (Rp)",
        "% Bahan",
        "% Ops",
        "% Profit",
        "Catatan",
    ]
    ws5.append(headers5)

    # Sheet 6: Riwayat Nota
    ws6 = wb.create_sheet("Riwayat Nota")
    headers6 = [
        "Tanggal",
        "Supplier",
        "No.Nota",
        "Item",
        "Qty",
        "Harga",
        "PPN",
        "Total",
        "Foto",
    ]
    ws6.append(headers6)

    # Apply bold header, freeze panes, and auto-width
    for ws in (ws1, ws2, ws3, ws4, ws5, ws6):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2

    out_path = base / "pembukuan_template.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()


