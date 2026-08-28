"""
backend/services/export_service.py
Export Excel Pembukuan MBG — Fixed version

Template sheets: Harian | Mingguan | Bulanan | Stok | Laporan Pemerintah | Riwayat Nota
Per user request: merge Harian+Mingguan+Bulanan → satu sheet "Bulanan"
Hasil akhir: Ringkasan | Bulanan | Stok | Laporan Pemerintah | Riwayat Nota

Upload ke Supabase Storage, upsert excel_files.
"""

import io
import logging
import os
from copy import copy
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

BULAN = {
    1: "Januari", 2: "Februari", 3: "Maret",     4: "April",
    5: "Mei",     6: "Juni",     7: "Juli",        8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "pembukuan_template.xlsx"

# ─── Konstanta styling (Dashboard Biru Modern v25) ─────────────────────────────
HDR_FILL_COLOR   = "0D47A1"  # Navy Blue
ALT_ROW_COLOR    = "F0F7FF"  # Very Light Blue (Zebra)
TOTAL_FILL_COLOR = "E3F2FD"  # Light Blue for Totals
BORDER_COLOR     = "D1D1D1"  # Soft Grey Border
GLOBAL_FONT      = "Segoe UI"
GLOBAL_FONT_SIZE = 11

# Accounting format: _($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)
ACCOUNTING_FORMAT = '"Rp "#,##0;[Red]"Rp "-#,##0;"-"'

# Minimum total daily portions to qualify for fixed insentif Rp 6jt.
# Days with fewer portions (e.g. test deliveries of 1-5 porsi) will NOT
# receive the fixed insentif to avoid extreme % anomalies in audit reports.
MIN_PORSI_FOR_INSENTIF = 50


def _d(v, default: Decimal = Decimal("0")) -> Decimal:
    try:
        return Decimal(str(v or 0))
    except Exception:
        return default


class ExportService:
    """Generate Excel pembukuan bulanan dan upload ke Supabase Storage."""

    # ─── Public: dipanggil dari BackgroundTasks atau endpoint ────────────────

    def regenerate_monthly_excel(
        self,
        tenant_id: str,
        year: int,
        month: int,
        supabase=None,
    ) -> Optional[str]:
        """
        Buat Excel untuk bulan year/month, upload ke Storage.
        Return: file_url (str) atau None jika gagal.
        """
        if supabase is None:
            from core.database import get_supabase
            supabase = get_supabase()

        from core.config import settings as cfg

        # Tandai sedang diproses
        period_key = f"{year:04d}-{month:02d}"
        self._upsert_excel_files(supabase, tenant_id, period_key, status="generating")

        try:
            # 1. Ambil semua data
            data = self._fetch_all(supabase, tenant_id, year, month, cfg)
            logger.info(
                f"=== GENERATE EXCEL {year}/{month} tenant={tenant_id} === "
                f"deliveries={len(data['deliveries'])}, "
                f"transactions={len(data['transactions'])}, "
                f"products={len(data['products'])}, "
                f"receivables={len(data['receivables'])}, "
                f"payables={len(data['payables'])}"
            )

            # 2. Buat workbook (BUKAN dari template — buat dari scratch agar pasti benar)
            wb = self._build_workbook(data, year, month)

            # 3. Save ke buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # 4. Upload ke Storage
            file_url = self._upload_to_storage(
                supabase, tenant_id, year, month, buffer.getvalue()
            )

            # 5. Upsert excel_files
            self._upsert_excel_files(
                supabase, tenant_id, period_key,
                status="ready", file_url=file_url
            )

            logger.info(f"Excel {period_key} tenant {tenant_id} selesai: {file_url}")
            return file_url

        except Exception as e:
            logger.error(f"regenerate_monthly_excel GAGAL: {e}", exc_info=True)
            self._upsert_excel_files(
                supabase, tenant_id, period_key,
                status="error", error_message=str(e)[:500]
            )
            return None

    # ─── Data fetching ────────────────────────────────────────────────────────

    def _fetch_all(self, supabase, tenant_id: str, year: int, month: int, cfg) -> Dict[str, Any]:
        first = date(year, month, 1)
        if month == 12:
            last = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last = date(year, month + 1, 1) - timedelta(days=1)

        logger.info(f"Fetching data for {year}-{month} (Range: {first} to {last}), tenant={tenant_id}")

        # Tenant info
        t_resp = (
            supabase.table("tenants")
            .select("id, name")
            .eq("id", tenant_id)
            .single()
            .execute()
        )
        tenant = getattr(t_resp, "data", None) or {"id": tenant_id, "name": "SPPG"}

        # Allocation settings
        alloc_resp = (
            supabase.table("mbg_allocation_settings")
            .select("*")
            .eq("tenant_id", tenant_id)
            .limit(1)
            .execute()
        )
        alloc_cfg = (getattr(alloc_resp, "data", None) or [{}])[0]
        price_per_portion = float(alloc_cfg.get("price_per_portion") or 15000)
        rate_bahan_sd = Decimal(str(alloc_cfg.get("bahan_sd_smp") or 10000))
        rate_bahan_tk = Decimal(str(alloc_cfg.get("bahan_paud_tk") or 8000))
        rate_ops_porsi = Decimal(str(alloc_cfg.get("ops_per_porsi") or 3000))
        insentif_harian = Decimal(str(alloc_cfg.get("insentif_harian") or 6000000))
        # New: Support for fixed monthly working days (default 27 as requested)
        working_days_month = int(alloc_cfg.get("hari_kerja_bulan") or 27)

        # MBG deliveries + schools
        del_resp = (
            supabase.table("mbg_deliveries")
            .select("*, schools(name, school_level, distance_km)")
            .eq("tenant_id", tenant_id)
            .gte("delivery_date", first.isoformat())
            .lte("delivery_date", last.isoformat())
            .order("delivery_date")
            .execute()
        )
        deliveries = getattr(del_resp, "data", None) or []

        # Weekly menus (for menu names AND nutrition)
        wm_resp = (
            supabase.table("mbg_weekly_menus")
            .select("menu_id, week_start, day_of_week, menu_name")
            .eq("tenant_id", tenant_id)
            .gte("week_start", (first - timedelta(days=7)).isoformat())
            .lte("week_start", last.isoformat())
            .execute()
        )
        weekly_menus = getattr(wm_resp, "data", None) or []
        
        # Recipes for nutrition calculation
        menu_ids = list(set([wm.get("menu_id") for wm in weekly_menus if wm.get("menu_id")]))
        recipes_all = []
        if menu_ids:
            try:
                r_resp = (
                    supabase.table("recipes")
                    .select("*, products!recipes_ingredient_id_fkey(id, name, unit, conversion_factor, nutrition_ref_id)")
                    .in_("menu_id", menu_ids)
                    .execute()
                )
                recipes_all = getattr(r_resp, "data", None) or []
            except Exception as e:
                logger.error(f"Failed to fetch recipes: {e}")
                
        # Nutrition Ref
        nutrition_ref_all = []
        try:
            n_resp = supabase.table("nutrition_ref").select("*").execute()
            nutrition_ref_all = getattr(n_resp, "data", None) or []
        except Exception as e:
            logger.error(f"Failed to fetch nutrition_ref: {e}")


        # Transactions (ALL confirmed, no type filter)
        trx_resp = (
            supabase.table("transactions")
            .select("id, date, suppliers(name), supplier_id, total, payment_method, status, nama_toko, juknis_category")
            .eq("tenant_id", tenant_id)
            .eq("status", "confirmed")
            .gte("date", first.isoformat())
            .lte("date", last.isoformat())
            .order("date")
            .execute()
        )
        transactions = getattr(trx_resp, "data", None) or []
        print(f"DEBUG: Found {len(transactions)} confirmed transactions")
        if transactions:
            print(f"DEBUG: Sample Transaction: {transactions[0]['date']} - {transactions[0].get('total')}")

        # Items per transaction
        trx_ids = [t["id"] for t in transactions]
        items_all: List[dict] = []
        if trx_ids:
            # Fetch in batches of 50 to avoid URL length issues
            for i in range(0, len(trx_ids), 50):
                batch = trx_ids[i:i+50]
                items_resp = (
                    supabase.table("transaction_items")
                    .select("*")
                    .in_("transaction_id", batch)
                    .execute()
                )
                items_all.extend(getattr(items_resp, "data", None) or [])

        # Supplier is_pkp status
        sup_ids = list({t.get("supplier_id") for t in transactions if t.get("supplier_id")})
        supplier_pkp: Dict[str, bool] = {}
        if sup_ids:
            sup_resp = (
                supabase.table("suppliers")
                .select("id, is_pkp")
                .in_("id", sup_ids)
                .execute()
            )
            supplier_pkp = {s["id"]: bool(s.get("is_pkp")) for s in (getattr(sup_resp, "data", None) or [])}

        # Products + stock history
        prod_resp = (
            supabase.table("products")
            .select("id, name, category, unit, display_unit, conversion_factor, harga, stock_qty, stock_min")
            .eq("tenant_id", tenant_id)
            .eq("is_active", True)
            .order("category")
            .order("name")
            .execute()
        )
        products = getattr(prod_resp, "data", None) or []

        hist_resp = (
            supabase.table("stock_history")
            .select("product_id, change_qty, reason")
            .eq("tenant_id", tenant_id)
            .gte("created_at", first.isoformat())
            .lte("created_at", (last + timedelta(days=1)).isoformat())
            .execute()
        )
        stock_history = getattr(hist_resp, "data", None) or []

        # Receivables
        recv_resp = (
            supabase.table("receivables")
            .select("*")
            .eq("tenant_id", tenant_id)
            .gte("created_at", first.isoformat())
            .lte("created_at", (last + timedelta(days=1)).isoformat())
            .order("created_at")
            .execute()
        )
        receivables = getattr(recv_resp, "data", None) or []

        # Payables
        pay_resp = (
            supabase.table("payables")
            .select("*")
            .eq("tenant_id", tenant_id)
            .gte("created_at", first.isoformat())
            .lte("created_at", (last + timedelta(days=1)).isoformat())
            .order("created_at")
            .execute()
        )
        payables = getattr(pay_resp, "data", None) or []

        # Operational Costs
        ops_resp = (
            supabase.table("operational_costs")
            .select("*, transactions(status)")
            .eq("tenant_id", tenant_id)
            .gte("cost_date", first.isoformat())
            .lt("cost_date", (last + timedelta(days=1)).isoformat())
            .order("cost_date")
            .execute()
        )
        # Filter only active (not voided) if we want, or show all
        operational_costs = getattr(ops_resp, "data", None) or []

        # Payroll Items
        period_resp = (
            supabase.table("payroll_periods")
            .select("id, name, start_date, status")
            .eq("tenant_id", tenant_id)
            .neq("status", "draft")
            .gte("start_date", first.isoformat())
            .lt("start_date", (last + timedelta(days=1)).isoformat())
            .execute()
        )
        periods = getattr(period_resp, "data", None) or []
        period_ids = [p["id"] for p in periods]
        
        payroll_items = []
        if period_ids:
            items_res = (
                supabase.table("payroll_items")
                .select("*, employees(name), payroll_periods(name, start_date, status)")
                .in_("period_id", period_ids)
                .execute()
            )
            payroll_items = getattr(items_res, "data", None) or []

        # Kas Ledger
        ledger_res = (
            supabase.table("kas_ledger")
            .select("*, kas_accounts(name, type)")
            .eq("tenant_id", tenant_id)
            .gte("entry_date", first.isoformat())
            .lte("entry_date", last.isoformat())
            .order("entry_date")
            .order("created_at")
            .execute()
        )
        kas_ledger = getattr(ledger_res, "data", None) or []

        # School Beneficiaries (for Sheet 9) — include mbg_allocation_type
        sb_res = (
            supabase.table("school_beneficiaries")
            .select("*, schools(name, mbg_allocation_type, default_portions, school_level), beneficiary_types(name)")
            .eq("tenant_id", tenant_id)
            .execute()
        )
        school_beneficiaries = getattr(sb_res, "data", None) or []

        # Fund Disbursements (for Buku Kas sheets)
        disb_res = (
            supabase.table("fund_disbursements")
            .select("*")
            .eq("tenant_id", tenant_id)
            .eq("year", year)
            .eq("month", month)
            .order("disbursement_date")
            .execute()
        )
        fund_disbursements = getattr(disb_res, "data", None) or []

        # Budget Summary
        from services.budget_service import budget_service
        budget_summary = budget_service.get_monthly_summary(tenant_id, year, month)

        # Budget Allocations (for exact per-portion reporting)
        ba_res = (
            supabase.table("mbg_budget_allocations")
            .select("*")
            .eq("tenant_id", tenant_id)
            .gte("date", first.isoformat())
            .lte("date", last.isoformat())
            .execute()
        )
        budget_allocations = getattr(ba_res, "data", None) or []

        # ─── Compliance data (Modul 21.5e) ──────────────────────────────────
        try:
            hyg_res = (
                supabase.table("hygiene_checks").select("*")
                .eq("tenant_id", tenant_id)
                .gte("check_date", first.isoformat())
                .lte("check_date", last.isoformat())
                .order("check_date")
                .execute()
            )
            hygiene_checks = getattr(hyg_res, "data", None) or []
        except Exception:
            hygiene_checks = []

        try:
            temp_res = (
                supabase.table("temperature_logs").select("*")
                .eq("tenant_id", tenant_id)
                .gte("log_date", first.isoformat())
                .lte("log_date", last.isoformat())
                .order("log_date")
                .execute()
            )
            temperature_logs = getattr(temp_res, "data", None) or []
        except Exception:
            temperature_logs = []

        try:
            smp_res = (
                supabase.table("food_samples").select("*")
                .eq("tenant_id", tenant_id)
                .eq("status", "disimpan")
                .order("expires_at")
                .execute()
            )
            food_samples = getattr(smp_res, "data", None) or []
        except Exception:
            food_samples = []

        try:
            waste_res = (
                supabase.table("food_waste_reports").select("*, schools(name)")
                .eq("tenant_id", tenant_id)
                .gte("report_date", first.isoformat())
                .lte("report_date", last.isoformat())
                .order("report_date")
                .execute()
            )
            food_waste_reports = getattr(waste_res, "data", None) or []
        except Exception:
            food_waste_reports = []

        try:
            inc_res = (
                supabase.table("incident_reports").select("*, schools(name)")
                .eq("tenant_id", tenant_id)
                .gte("incident_time", f"{first}T00:00:00")
                .lte("incident_time", f"{last}T23:59:59")
                .order("incident_time", desc=True)
                .execute()
            )
            incident_reports = getattr(inc_res, "data", None) or []
        except Exception:
            incident_reports = []

        # ── Fixed Fee Model: Rp 6jt/hari × hari kerja (dari settings) ──
        total_portions_all = sum(_d(d.get("portions_sent", 0)) for d in deliveries)
        total_profit_target = insentif_harian * working_days_month
        # Count unique delivery days
        delivery_days = len(set(d.get("delivery_date", "") for d in deliveries if d.get("delivery_date")))
        # Fee per porsi rounded to INTEGER (no decimals for audit)
        fee_per_porsi = Decimal("0")
        if total_portions_all > 0:
            fee_per_porsi = (total_profit_target / total_portions_all).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

        return {
            "tenant": tenant,
            "first": first,
            "last": last,
            "price_per_portion": price_per_portion,
            "rate_bahan_sd": rate_bahan_sd,
            "rate_bahan_tk": rate_bahan_tk,
            "rate_ops_porsi": rate_ops_porsi,
            "insentif_harian": insentif_harian,
            "working_days_month": working_days_month,
            "total_portions_all": total_portions_all,
            "total_profit_target": total_profit_target,
            "fee_per_porsi": fee_per_porsi,
            "delivery_days": delivery_days,
            "deliveries": deliveries,
            "weekly_menus": weekly_menus,
            "recipes_all": recipes_all,
            "nutrition_ref_all": nutrition_ref_all,
            "transactions": transactions,
            "items_all": items_all,
            "supplier_pkp": supplier_pkp,
            "products": products,
            "stock_history": stock_history,
            "receivables": receivables,
            "payables": payables,
            "operational_costs": operational_costs,
            "payroll_items": payroll_items,
            "kas_ledger": kas_ledger,
            "school_beneficiaries": school_beneficiaries,
            "budget_summary": budget_summary,
            "budget_allocations": budget_allocations,
            "master_schedule": self._fetch_schedule_for_export(tenant_id),
            "fund_disbursements": fund_disbursements,
            "hygiene_checks": hygiene_checks,
            "temperature_logs": temperature_logs,
            "food_samples": food_samples,
            "food_waste_reports": food_waste_reports,
            "incident_reports": incident_reports,
        }

    # ─── Workbook builder ─────────────────────────────────────────────────────

    def _build_workbook(self, data: dict, year: int, month: int):
        from openpyxl import Workbook
        wb = Workbook()

        # Sheet 1: Ringkasan (summary)
        ws_ringkasan = wb.active
        ws_ringkasan.title = "Ringkasan"

        # Sheet 2: Bulanan (merged harian+mingguan+bulanan — all transactions)
        ws_bulanan = wb.create_sheet("Bulanan")

        # Sheet 3: Stok
        ws_stok = wb.create_sheet("Stok")

        # Sheet 4: Laporan Pemerintah
        ws_pemerintah = wb.create_sheet("Laporan Pemerintah")

        # Sheet 5: Riwayat Nota
        ws_nota = wb.create_sheet("Riwayat Nota")
        
        # Sheet 6: Operasional & Upah
        ws_ops = wb.create_sheet("Operasional & Upah")

        # Sheet 7: Buku Kas Umum
        ws_bku = wb.create_sheet("Buku Kas Umum")

        # Sheet 8: Buku Kas Kecil
        ws_bkk = wb.create_sheet("Buku Kas Kecil")

        # Sheet 9: Nutrisi & Penerima Manfaat
        ws_pm = wb.create_sheet("Nutrisi & Penerima Manfaat")

        # Sheet 10: Rekapitulasi Juknis
        ws_rj = wb.create_sheet("Rekapitulasi Juknis")

        # Sheet 11-15: Compliance (Modul 21.5e)
        ws_hygiene = wb.create_sheet("Higiene")
        ws_suhu = wb.create_sheet("Suhu")
        ws_sampel = wb.create_sheet("Bank Sampel")
        ws_sisa = wb.create_sheet("Sisa Makanan")
        ws_insiden = wb.create_sheet("Insiden")

        logger.info(f"Building workbook with sheets: {wb.sheetnames}")

        # Track per-sheet errors so one bad sheet doesn't kill everything
        sheet_errors = []

        def _safe_fill(sheet_name, fill_fn, *args, **kwargs):
            try:
                fill_fn(*args, **kwargs)
                logger.info(f"  ✅ {sheet_name} OK")
            except Exception as e:
                logger.error(f"  ❌ {sheet_name} GAGAL: {e}", exc_info=True)
                sheet_errors.append(f"{sheet_name}: {str(e)[:200]}")

        _safe_fill("Ringkasan", self._fill_ringkasan, ws_ringkasan, data, year, month)
        _safe_fill("Bulanan", self._fill_bulanan, ws_bulanan, data, year, month)
        _safe_fill("Stok", self._fill_stok, ws_stok, data)
        _safe_fill("Laporan Pemerintah", self._fill_laporan_pemerintah, ws_pemerintah, data, year, month)
        _safe_fill("Riwayat Nota", self._fill_riwayat_nota, ws_nota, data)
        _safe_fill("Operasional & Upah", self._fill_ops_upah, ws_ops, data, year, month)
        _safe_fill("Buku Kas Umum", self._fill_bku, ws_bku, data, year, month, "va_bank", "BUKU KAS UMUM")
        _safe_fill("Buku Kas Kecil", self._fill_bku, ws_bkk, data, year, month, "kas_kecil", "BUKU KAS KECIL")
        _safe_fill("Nutrisi & PM", self._fill_penerima_manfaat, ws_pm, data, year, month)
        _safe_fill("Rekapitulasi Juknis", self._fill_rekapitulasi_juknis, ws_rj, data, year, month)

        # Fill compliance sheets
        _safe_fill("Higiene", self._fill_hygiene, ws_hygiene, data, year, month)
        _safe_fill("Suhu", self._fill_suhu, ws_suhu, data, year, month)
        _safe_fill("Bank Sampel", self._fill_bank_sampel, ws_sampel, data)
        _safe_fill("Sisa Makanan", self._fill_sisa_makanan, ws_sisa, data, year, month)
        _safe_fill("Insiden", self._fill_insiden, ws_insiden, data, year, month)

        if sheet_errors:
            logger.warning(f"Excel generated with {len(sheet_errors)} sheet error(s): {sheet_errors}")

        return wb

    # ─── Sheet 1: Ringkasan ───────────────────────────────────────────────────

    def _fill_bku(self, ws, data: dict, year: int, month: int, kas_type: str, title: str) -> None:
        logger.info(f"Filling {title} (Kategori: {kas_type})...")
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]
        
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{title} — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["No", "Tanggal", "Akun Kas", "Keterangan", "Masuk (Debit)", "Keluar (Kredit)", "Saldo Berjalan"]
        self._write_headers(ws, 3, headers)

        # Group portions by day for virtual allocation
        portions_by_day = {}
        for d in data.get("deliveries", []):
            dt = d.get("delivery_date", "")
            if dt:
                portions_by_day[dt] = portions_by_day.get(dt, 0) + d.get("portions_sent", 0)

        # Filter ledger by kas_type
        filtered_ledger = []
        for e in data.get("kas_ledger", []):
            acc_info = e.get("kas_accounts")
            if isinstance(acc_info, dict) and acc_info.get("type") == kas_type:
                filtered_ledger.append(e)

        # Build list of entries to display
        entries_to_show = []
        
        if filtered_ledger:
            # --- Use real ledger entries ---
            for e in filtered_ledger:
                is_debit = e.get("entry_type") == "debit"
                acc_info = e.get("kas_accounts") or {}
                entries_to_show.append({
                    "date": e.get("entry_date", ""),
                    "acc": acc_info.get("name", title),
                    "desc": e.get("description", ""),
                    "masuk": float(e.get("amount", 0)) if is_debit else 0,
                    "keluar": 0 if is_debit else float(e.get("amount", 0)),
                    "sort_key": (e.get("entry_date", ""), 0 if "saldo" in (e.get("description", "") or "").lower() else 1, e.get("created_at", ""))
                })
        else:
            # --- Synthesize entries if ledger is empty ---
            # Debits from disbursements
            if kas_type == "va_bank":
                for disb in data.get("fund_disbursements", []):
                    ref = disb.get("reference_number", "")
                    entries_to_show.append({
                        "date": disb.get("disbursement_date", ""),
                        "acc": title,
                        "desc": f"Pencairan Dana MBG (Ref: {ref})" if ref else "Pencairan Dana MBG",
                        "masuk": float(_d(disb.get("amount"))),
                        "keluar": 0,
                        "sort_key": (disb.get("disbursement_date", ""), 0, "A")
                    })
            # Credits from transactions
            for trx in data.get("transactions", []):
                if trx.get("status") in ("void", "cancelled"): continue
                supplier = trx.get("nama_toko", "") or "Supplier"
                entries_to_show.append({
                    "date": trx.get("date", ""),
                    "acc": title,
                    "desc": f"Pembelian bahan - {supplier}",
                    "masuk": 0,
                    "keluar": float(_d(trx.get("total"))),
                    "sort_key": (trx.get("date", ""), 1, trx.get("id", ""))
                })

        # --- Inject Virtual Allocation (Fixed Model) for va_bank ---
        if kas_type == "va_bank":
            insentif_harian = float(data.get("insentif_harian", 6000000))
            rate_ops = float(data.get("rate_ops_porsi", 3000))
            for dt, porsi in portions_by_day.items():
                allocation = insentif_harian + (porsi * rate_ops)
                entries_to_show.append({
                    "date": dt,
                    "acc": title,
                    "desc": "Alokasi Management Fee & Operasional SPPG (Fixed)",
                    "masuk": 0,
                    "keluar": allocation,
                    "sort_key": (dt, 2, "VIRTUAL") # After real/purchase entries
                })

        # Sort all entries
        entries_to_show.sort(key=lambda x: x["sort_key"])

        row = 4
        n = 0
        running_balance = Decimal("0")
        
        for e in entries_to_show:
            n += 1
            running_balance += Decimal(str(e["masuk"]))
            running_balance -= Decimal(str(e["keluar"]))
            
            ws.cell(row, 1, n)
            ws.cell(row, 2, e["date"])
            ws.cell(row, 3, e["acc"])
            ws.cell(row, 4, e["desc"])
            ws.cell(row, 5, e["masuk"]).number_format = '"Rp "#,##0'
            ws.cell(row, 6, e["keluar"]).number_format = '"Rp "#,##0'
            ws.cell(row, 7, float(running_balance)).number_format = '"Rp "#,##0'
            
            self._alt_row(ws, row, 7)
            row += 1

        if n == 0:
            ws.cell(row, 1, "—")
            ws.cell(row, 2, f"Tidak ada transaksi pada {title} bulan ini")
            row += 1

        # Total row
        if n > 0:
            row += 1
            total_debit = Decimal("0")
            total_kredit = Decimal("0")
            if filtered_ledger:
                for e in filtered_ledger:
                    if e.get("entry_type") == "debit":
                        total_debit += _d(e.get("amount"))
                    else:
                        total_kredit += _d(e.get("amount"))
            else:
                for disb in data.get("fund_disbursements", []):
                    total_debit += _d(disb.get("amount"))
                for trx in data.get("transactions", []):
                    if trx.get("status") not in ("void", "cancelled"):
                        total_kredit += _d(trx.get("total"))

            self._write_total_row(ws, row, {
                5: float(total_debit), 
                6: float(total_kredit),
                7: float(total_debit - total_kredit)
            }, max_col=7)

        self._set_column_widths(ws)
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:G{row if n > 0 else 3}"

    def _fill_penerima_manfaat(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (9/10): Filling Nutrisi & Penerima Manfaat...")
        from openpyxl.styles import Font, Alignment, PatternFill
        from calendar import monthrange
        from routers.nutrition import _recipe_to_grams
        import json
        
        bulan_str = BULAN[month]
        
        # --- TABEL 1: Rekap Nutrisi Bulanan ---
        ws.merge_cells("A1:P1")
        ws["A1"] = f"REKAP NUTRISI BULANAN — LAPORAN PEMERINTAH ({bulan_str} {year})"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")
        
        headers_nutrisi = [
            "Tgl", "Hari", "Menu", "Porsi", "Kalori (kkal)", "Protein (g)", 
            "Lemak (g)", "Karbo (g)", "Kalsium (mg)", "Zat Besi (mg)", 
            "Vit C (mg)", "Seng (mg)", "Vit A (mcg)", "Berat (g)", "Sayur %", "Status"
        ]
        self._write_headers(ws, 3, headers_nutrisi)
        
        _, days_in_month = monthrange(year, month)
        first_day = date(year, month, 1)
        last_day = date(year, month, days_in_month)
        
        # Map menus by date
        menu_by_date = {}
        for wm in data.get("weekly_menus", []):
            try:
                ws_date = date.fromisoformat(wm["week_start"])
                dow = wm.get("day_of_week", 1)
                actual_date = ws_date + timedelta(days=dow - 1)
                if first_day <= actual_date <= last_day:
                    menu_by_date[actual_date.isoformat()] = {
                        "menu_id": wm.get("menu_id"),
                        "menu_name": wm.get("menu_name", "")
                    }
            except Exception:
                pass
                
        # Group deliveries
        deliveries_by_date = {}
        for d in data.get("deliveries", []):
            dt = d.get("delivery_date")
            if dt:
                deliveries_by_date[dt] = deliveries_by_date.get(dt, 0) + d.get("portions_sent", 0)
                
        # Nut cache
        recipes_all = data.get("recipes_all", [])
        nut_refs = {n["id"]: n for n in data.get("nutrition_ref_all", [])}
        
        nut_cache = {}
        for r in recipes_all:
            mid = r.get("menu_id")
            if mid not in nut_cache:
                nut_cache[mid] = []
            nut_cache[mid].append(r)
            
        def get_nut_totals(mid):
            if not mid or mid not in nut_cache:
                return None
            recs = nut_cache[mid]
            t_cal=0.0; t_pro=0.0; t_fat=0.0; t_carb=0.0; t_gram=0.0; s_gram=0.0
            t_ca=0.0; t_fe=0.0; t_vitc=0.0; t_zn=0.0; t_vita=0.0
            
            for r in recs:
                if r.get("usage_type", "per_porsi") != "per_porsi": continue
                prod = r.get("products")
                if not prod: continue
                weight = _recipe_to_grams(r, prod)
                nut_id = prod.get("nutrition_ref_id")
                nref = nut_refs.get(nut_id) if nut_id else None
                kat = ""
                if nref and weight > 0:
                    m = weight / 100.0
                    t_cal += float(nref.get("calories", 0) or 0) * m
                    t_pro += float(nref.get("proteins", 0) or 0) * m
                    t_fat += float(nref.get("fat", 0) or 0) * m
                    t_carb += float(nref.get("carbohydrate", 0) or 0) * m
                    
                    # Accumulate micro-nutrients
                    cust = nref.get("custom_nutrients") or {}
                    if isinstance(cust, str):
                        try:
                            cust = json.loads(cust)
                        except Exception:
                            cust = {}
                    t_ca += float(cust.get("kalsium_mg", 0) or 0) * m
                    t_fe += float(cust.get("besi_mg", 0) or 0) * m
                    t_vitc += float(cust.get("vitamin_c_mg", 0) or 0) * m
                    t_zn += float(cust.get("seng_mg", 0) or 0) * m
                    # Retinol + Beta Karoten / Karoten Total
                    t_vita += float(cust.get("retinol_mcg", 0) or cust.get("karoten_total_mcg", 0) or 0) * m
                    
                    kat = nref.get("kategori", "")
                t_gram += weight
                if "sayur" in kat.lower():
                    s_gram += weight
            spct = round((s_gram / t_gram)*100, 1) if t_gram > 0 else 0.0
            return {
                "cal": round(t_cal), "pro": round(t_pro, 1), "fat": round(t_fat, 1), 
                "carb": round(t_carb, 1), "gram": round(t_gram), "spct": spct,
                "balanced": spct >= 30.0,
                # Micros
                "ca": round(t_ca, 1),
                "fe": round(t_fe, 2),
                "vitc": round(t_vitc, 1),
                "zn": round(t_zn, 2),
                "vita": round(t_vita, 1)
            }
        
        HARI_INDO = {0: "Senin", 1: "Selasa", 2: "Rabu", 3: "Kamis", 4: "Jumat", 5: "Sabtu", 6: "Minggu"}
        
        row = 4
        tot_porsi = 0
        from collections import defaultdict
        sums = defaultdict(float)
        days_with_nut = 0
        
        for day_num in range(1, days_in_month + 1):
            curr = date(year, month, day_num)
            d_str = curr.isoformat()
            wd = curr.weekday()
            
            if wd == 6: continue # Skip sunday
            
            minfo = menu_by_date.get(d_str, {})
            mname = minfo.get("menu_name", "")
            mid = minfo.get("menu_id")
            
            nt = get_nut_totals(mid)
            porsi = deliveries_by_date.get(d_str, 0)
            tot_porsi += porsi
            
            ws.cell(row, 1, day_num)
            ws.cell(row, 2, HARI_INDO.get(wd, ""))
            ws.cell(row, 3, mname or "—")
            ws.cell(row, 4, porsi or "—")
            
            if nt:
                ws.cell(row, 5, nt["cal"])
                ws.cell(row, 6, nt["pro"])
                ws.cell(row, 7, nt["fat"])
                ws.cell(row, 8, nt["carb"])
                ws.cell(row, 9, nt["ca"])
                ws.cell(row, 10, nt["fe"])
                ws.cell(row, 11, nt["vitc"])
                ws.cell(row, 12, nt["zn"])
                ws.cell(row, 13, nt["vita"])
                ws.cell(row, 14, nt["gram"])
                ws.cell(row, 15, f'{nt["spct"]}%')
                ws.cell(row, 16, "✅" if nt["balanced"] else "⚠️")
                
                sums["cal"] += nt["cal"]
                sums["pro"] += nt["pro"]
                sums["fat"] += nt["fat"]
                sums["carb"] += nt["carb"]
                sums["ca"] += nt["ca"]
                sums["fe"] += nt["fe"]
                sums["vitc"] += nt["vitc"]
                sums["zn"] += nt["zn"]
                sums["vita"] += nt["vita"]
                sums["gram"] += nt["gram"]
                sums["spct"] += nt["spct"]
                if nt["balanced"]: sums["bal"] += 1
                days_with_nut += 1
            else:
                for c in range(5, 17):
                    ws.cell(row, c, "—")
            
            self._alt_row(ws, row, 16)
            row += 1
            
        # Rata-rata / Total row
        row += 1
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row, 1, "RATA-RATA / TOTAL").font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="right")
        ws.cell(row, 4, tot_porsi).font = Font(bold=True)
        
        if days_with_nut > 0:
            ws.cell(row, 5, round(sums["cal"] / days_with_nut)).font = Font(bold=True)
            ws.cell(row, 6, round(sums["pro"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 7, round(sums["fat"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 8, round(sums["carb"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 9, round(sums["ca"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 10, round(sums["fe"] / days_with_nut, 2)).font = Font(bold=True)
            ws.cell(row, 11, round(sums["vitc"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 12, round(sums["zn"] / days_with_nut, 2)).font = Font(bold=True)
            ws.cell(row, 13, round(sums["vita"] / days_with_nut, 1)).font = Font(bold=True)
            ws.cell(row, 14, round(sums["gram"] / days_with_nut)).font = Font(bold=True)
            ws.cell(row, 15, f'{round(sums["spct"] / days_with_nut, 1)}%').font = Font(bold=True)
            ws.cell(row, 16, f'{int(sums["bal"])}/{days_with_nut}').font = Font(bold=True)
        
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=TOTAL_FILL_COLOR)
        for c in range(4, 17):
            ws.cell(row, c).fill = PatternFill("solid", fgColor=TOTAL_FILL_COLOR)
        
        row += 3
        
        # --- TABEL 2: Data Penerima Manfaat ---
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, f"DATA PENERIMA MANFAAT — {bulan_str} {year}")
        ws.cell(row, 1).font = Font(name=GLOBAL_FONT, bold=True, size=14, color="FFFFFF")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        
        row += 2
        headers_pm = ["No", "Sekolah", "Tipe Alokasi", "Jenis Penerima", "Porsi Target/Hari", "Total Terkirim"]
        self._write_headers(ws, row, headers_pm)
        row += 1
        
        sb_data = data.get("school_beneficiaries", [])
        
        # Calculate delivered per school manually since deliveries data is flat
        delivered_by_school = {}
        for d in data.get("deliveries", []):
            sid = d.get("school_id")
            if sid:
                delivered_by_school.setdefault(sid, 0)
                delivered_by_school[sid] += d.get("portions_sent", 0)
                
        tot_target = 0
        tot_sent = 0
        
        if not sb_data:
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Data Penerima Kosong")
            row += 1
        else:
            for i, sb in enumerate(sb_data, 1):
                sch = sb.get("schools") or {}
                sname = sch.get("name", "")
                sid = sb.get("school_id", "")
                bt_name = (sb.get("beneficiary_types") or {}).get("name", "")
                slevel = sch.get("school_level", "sd_smp")
                
                # Check column mbg_allocation_type if exists, else fallback to level
                alloc_type = sch.get("mbg_allocation_type", "")
                if not alloc_type:
                    alloc_type = "paud_tk" if slevel == "paud_tk" else "sd_smp"
                al_label = "Rp 8.000 (PAUD/TK)" if alloc_type == "paud_tk" else "Rp 10.000 (SD/SMP/SMA)"
                if alloc_type == "guru": al_label = "Rp 15.000 (Guru)"
                
                target_porsi = sb.get("target_portions", 0) or sch.get("default_portions", 0)
                sent = delivered_by_school.get(sid, 0)
                
                tot_target += target_porsi
                tot_sent += sent
                
                ws.cell(row, 1, i)
                ws.cell(row, 2, sname)
                ws.cell(row, 3, al_label)
                ws.cell(row, 4, bt_name)
                ws.cell(row, 5, target_porsi)
                ws.cell(row, 6, sent)
                
                self._alt_row(ws, row, 6)
                row += 1
                
        # Total Row PM
        self._write_total_row(ws, row, {5: tot_target, 6: tot_sent}, max_col=6, currency_cols=[])
        
        self._set_column_widths(ws)

    def _fetch_schedule_for_export(self, tenant_id: str) -> dict:
        """Fetch master schedule data for Excel export."""
        try:
            from routers.schedules import _fetch_master_data
            return _fetch_master_data(tenant_id) or {}
        except Exception as e:
            logger.warning(f"Failed to fetch master schedule for export: {e}")
            return {}

    def _fill_rekapitulasi_juknis(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (10/10): Filling Rekapitulasi Juknis (Audit Alignment)...")
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]
        
        ws.merge_cells("A1:E1")
        ws["A1"] = f"REKAPITULASI ANGGARAN (PER PORSI) — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        
        headers = ["Keterangan", "Anggaran (Target)", "Realisasi (Aktual)", "Selisih", "%"]
        self._write_headers(ws, 3, headers)
        
        deliveries = data.get("deliveries", [])
        transactions = data.get("transactions", [])
        budget = data.get("budget_summary", {})
        
        # Dana masuk dari pencairan
        total_disbursed = Decimal(budget.get("total_disbursed", 0))
        
        # Realisasi: Bahan = ONLY 'bahan_pangan' category
        exp_bahan = Decimal("0")
        ops_from_trx = Decimal("0")
        for t in transactions:
            cat = str(t.get("juknis_category") or "").lower()
            if cat == "bahan_pangan":
                exp_bahan += _d(t.get("total"))
            else:
                ops_from_trx += _d(t.get("total"))

        # Operasional = Ops costs + Gaji + Nota non-bahan (semua dari Rp 3k/porsi)
        exp_gaji = sum(_d(item.get("net_amount")) for item in data.get("payroll_items", []))
        exp_ops_riil = sum(_d(cost.get("amount")) for cost in data.get("operational_costs", []))
        exp_ops = exp_ops_riil + exp_gaji + ops_from_trx
        total_exp = exp_bahan + exp_ops
        
        # Target dari settings (SEWA DAPUR MODEL)
        rate_bahan_sd = data.get("rate_bahan_sd", Decimal("10000"))
        rate_bahan_tk = data.get("rate_bahan_tk", Decimal("8000"))
        rate_ops_p = data.get("rate_ops_porsi", Decimal("3000"))
        insentif_harian = data.get("insentif_harian", Decimal("6000000"))
        working_days = data.get("working_days_month", 26)
        
        target_bahan = Decimal("0")
        for d in deliveries:
            p = d.get("portions_sent", 0)
            s_data = (d.get("schools") or [{}])[0] if isinstance(d.get("schools"), list) else (d.get("schools") or {})
            sl = s_data.get("school_level", "sd_smp")
            target_bahan += Decimal(p) * (rate_bahan_tk if sl == "paud_tk" else rate_bahan_sd)
        
        total_porsi = data.get("total_portions_all", 0)
        target_ops = Decimal(total_porsi) * rate_ops_p   # Rp 3k/porsi = ops + gaji
        sewa_dapur = insentif_harian * working_days       # Profit pemilik
        
        # Cadangan = sisa efisiensi bahan (jika ada)
        cadangan_bahan = max(target_bahan - exp_bahan, Decimal("0"))
        
        target_total = target_bahan + target_ops
        total_pagu = target_total + sewa_dapur
        sisa_anggaran = total_disbursed - total_exp - sewa_dapur
        
        def pct(aktual, target):
            return f"{round(float(aktual/target*100), 2):.2f}%" if target > 0 else "0.00%"
            
        rows_data = [
            ("1. Pagu Anggaran Target", total_pagu, total_pagu, 0, "100%"),
            ("2. Pagu Yang Sudah Cair", total_disbursed, total_disbursed, 0, "100%"),
            ("-", "-", "-", "-", "-"),
            ("3. Bahan per Porsi", target_bahan, exp_bahan, target_bahan - exp_bahan, pct(exp_bahan, target_bahan)),
            ("4. Operasional", target_ops, exp_ops, target_ops - exp_ops, pct(exp_ops, target_ops)),
            ("-", "-", "-", "-", "-"),
            ("5. Total Pengeluaran (Bahan per porsi + Operasional)", target_total, total_exp, target_total - total_exp, pct(total_exp, target_total)),
            ("6. Sewa Dapur", sewa_dapur, sewa_dapur, 0, "100%"),
            ("-", "-", "-", "-", "-"),
            ("7. Sisa Anggaran (Pagu Cair - Total Pengeluaran - Sewa Dapur)", sisa_anggaran, sisa_anggaran, 0, ""),
        ]
        
        row = 4
        for rd in rows_data:
            ws.cell(row, 1, rd[0])
            if rd[0] == "-":
                row += 1
                continue
            
            # Values
            ws.cell(row, 2, float(rd[1])).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 3, float(rd[2])).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 4, float(rd[3])).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 5, rd[4])
            
            # Styling
            self._alt_row(ws, row, 5)
            
            # Highlight Total Keseluruhan
            if rd[0] == "Total Keseluruhan":
                from openpyxl.styles import PatternFill, Font
                highlight = PatternFill("solid", fgColor="BBDEFB") # Light Blue
                for c in range(1, 6):
                    ws.cell(row, c).fill = highlight
                    ws.cell(row, c).font = Font(name=GLOBAL_FONT, bold=True, size=12)
            
            row += 1
            
        self._set_column_widths(ws)
        ws.freeze_panes = "A4"

    def _fill_ringkasan(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (1/10): Filling Ringkasan Dashboard...")
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        bulan_str = BULAN[month]
        tenant_name = data["tenant"].get("name", "SPPG")

        # Global Font
        ws.sheet_view.showGridLines = False
        default_font = Font(name=GLOBAL_FONT, size=GLOBAL_FONT_SIZE)

        # Header Title
        ws.merge_cells("A1:C1")
        ws["A1"] = f"PEMBUKUAN CATERING MBG — {bulan_str} {year}"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Info Section
        ws["A3"] = "Nama SPPG:"; ws["A3"].font = Font(name=GLOBAL_FONT, bold=True)
        ws["B3"] = tenant_name;   ws["B3"].font = default_font
        ws["A4"] = "Periode:";    ws["A4"].font = Font(name=GLOBAL_FONT, bold=True)
        ws["B4"] = f"{data['first'].strftime('%d/%m/%Y')} s/d {data['last'].strftime('%d/%m/%Y')}"
        ws["B4"].font = default_font

        # Calculations
        deliveries = data["deliveries"]
        transactions = data["transactions"]
        budget = data.get("budget_summary", {})
        total_portions = sum(d.get("portions_sent", 0) for d in deliveries)
        total_days = len({d["delivery_date"] for d in deliveries})
        
        pagu_amount = Decimal(budget.get("pagu_amount", 0))
        total_disbursed = Decimal(budget.get("total_disbursed", 0))
        sisa_anggaran = Decimal(budget.get("sisa_anggaran", 0))

        revenue_net = total_disbursed
        
        # Calculate realisasi: Bahan = ONLY 'bahan_pangan' category
        # Anything else (uncategorized/other) moves to Operasional
        exp_bahan = Decimal("0")
        ops_from_trx = Decimal("0")
        for t in transactions:
            cat = str(t.get("juknis_category") or "").lower()
            if cat == "bahan_pangan":
                exp_bahan += _d(t.get("total"))
            else:
                ops_from_trx += _d(t.get("total"))

        # Operasional = Operational Costs + Gaji + Nota Non-Bahan (semua dari Rp 3k/porsi)
        exp_ops_riil = sum(_d(cost.get("amount")) for cost in data.get("operational_costs", []))
        exp_gaji = sum(_d(item.get("net_amount")) for item in data.get("payroll_items", []))
        exp_ops = exp_ops_riil + exp_gaji + ops_from_trx
        
        # Sewa Dapur = Rp 6jt/hari × hari kerja (keuntungan bersih pemilik)
        working_days = data.get("working_days_month", 26)
        insentif_harian = data.get("insentif_harian", Decimal("6000000"))
        sewa_dapur = insentif_harian * working_days
        
        total_exp = exp_bahan + exp_ops
        gross_profit = revenue_net - total_exp - sewa_dapur

        # Layout Tables
        sections = [
            ("ANGGARAN & PENDAPATAN", [
                ("Pagu Anggaran", float(pagu_amount)),
                ("Dana Masuk / Cair", float(total_disbursed)),
                ("Sisa Anggaran", float(sisa_anggaran)),
            ]),
            ("OPERASIONAL (VOLUME)", [
                ("Total Porsi Terkirim", f"{total_portions:,} porsi"),
                ("Total Hari Kirim", f"{total_days} hari"),
                ("Total Transaksi (Nota)", f"{len(transactions)} nota"),
            ]),
            ("REALISASI PENGELUARAN", [
                ("Bahan Pangan", float(exp_bahan)),
                ("Operasional + Gaji (Rp 3.000/porsi)", float(exp_ops)),
                ("TOTAL PENGELUARAN", float(total_exp)),
            ]),
            ("SEWA DAPUR", [
                ("Keuntungan Pemilik", float(sewa_dapur)),
            ]),
            ("RINGKASAN AKHIR", [
                ("SISA DANA", float(gross_profit)),
            ])
        ]

        row = 6
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        lbl_font = Font(name=GLOBAL_FONT, bold=True)
        val_font = Font(name=GLOBAL_FONT, color="000000", bold=True)
        border_all = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )

        for sec_title, items in sections:
            # Section Header
            ws.cell(row, 1, sec_title).font = Font(name=GLOBAL_FONT, bold=True, color="0D47A1")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row, 1).border = border_all
            ws.cell(row, 2).border = border_all
            row += 1
            
            for label, val in items:
                lbl_cell = ws.cell(row, 1, label)
                lbl_cell.font = lbl_font
                lbl_cell.border = border_all
                
                val_cell = ws.cell(row, 2, val)
                val_cell.font = val_font
                val_cell.border = border_all
                if isinstance(val, (float, Decimal, int)) and "porsi" not in str(val):
                    val_cell.number_format = ACCOUNTING_FORMAT
                
                row += 1
            row += 1 # Spacer

        # ─── Analisis Efisiensi & Surplus (Dashboard Dua Lini) ───
        ws.cell(row, 1, "ANALISIS EFISIENSI & SURPLUS (DASHBOARD INTERNAL)").font = Font(name=GLOBAL_FONT, bold=True, color="0D47A1")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        
        headers_dashboard = ["Sektor Perbandingan", "Target (Pagu/Piutang)", "Realisasi (Riil)", "Selisih (Surplus)", "Status"]
        for i, h in enumerate(headers_dashboard, 1):
            cell = ws.cell(row, i, h)
            cell.font = Font(name=GLOBAL_FONT, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_all
        row += 1

        # ── Logic for Pagu (Target) — Model Sewa Dapur ──
        total_portions_all = data.get("total_portions_all", 0)
        rate_bahan_sd = data.get("rate_bahan_sd", Decimal("10000"))
        rate_bahan_tk = data.get("rate_bahan_tk", Decimal("8000"))
        rate_ops = data.get("rate_ops_porsi", Decimal("3000"))
        
        # Pagu Bahan from deliveries
        pagu_bahan = Decimal("0")
        for d in data.get("deliveries", []):
            p = _d(d.get("portions_sent", 0))
            s_data = (d.get("schools") or [{}])[0] if isinstance(d.get("schools"), list) else (d.get("schools") or {})
            sl = s_data.get("school_level", "sd_smp")
            pagu_bahan += p * (rate_bahan_tk if sl == "paud_tk" else rate_bahan_sd)
        
        # Pagu Ops = Rp 3k/porsi (sumber dana BBM, gas, logistik DAN gaji)
        pagu_ops = _d(total_portions_all) * rate_ops
        
        # ── Riil (Realisasi) ──
        # Bahan: ONLY bahan_pangan category
        riil_bahan = sum(_d(t.get("total")) for t in data.get("transactions", []) 
                        if str(t.get("juknis_category") or "").lower() == "bahan_pangan")
        # Operasional: ops costs + gaji + nota non-bahan (semua dari Rp 3k/porsi)
        riil_ops_base = sum(_d(cost.get("amount")) for cost in data.get("operational_costs", []))
        riil_gaji = sum(_d(item.get("net_amount")) for item in data.get("payroll_items", []))
        riil_ops_trx = sum(_d(t.get("total")) for t in data.get("transactions", []) 
                         if str(t.get("juknis_category") or "").lower() != "bahan_pangan")
        riil_ops = riil_ops_base + riil_gaji + riil_ops_trx
        
        # ── Dashboard 2 Lini (Pengeluaran) + Sewa Dapur (Profit) ──
        cadangan_bahan = max(pagu_bahan - riil_bahan, Decimal("0"))
        
        dashboard_rows = [
            ("1. Bahan Baku (Pangan)", pagu_bahan, riil_bahan),
            ("2. Operasional + Gaji (BBM, Gas, Logistik, Karyawan)", pagu_ops, riil_ops),
        ]
        
        total_pagu_all = pagu_bahan + pagu_ops
        total_riil_all = riil_bahan + riil_ops
        
        for name, pagu, riil in dashboard_rows:
            diff = pagu - riil
            status = "✅ Hemat" if diff >= 0 else "⚠️ Over"
            ws.cell(row, 1, name).border = border_all
            ws.cell(row, 2, float(pagu)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 2).border = border_all
            ws.cell(row, 3, float(riil)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 3).border = border_all
            ws.cell(row, 4, float(diff)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 4).border = border_all
            if diff < 0:
                ws.cell(row, 4).font = Font(bold=True, color="FF0000")
            ws.cell(row, 5, status).border = border_all
            ws.cell(row, 5).alignment = Alignment(horizontal="center")
            row += 1
        
        # Cadangan Dana (Efisiensi Bahan)
        if cadangan_bahan > 0:
            ws.cell(row, 1, "💰 Cadangan Biaya Operasional (Efisiensi Bahan)").border = border_all
            ws.cell(row, 1).font = Font(italic=True, color="006100")
            ws.cell(row, 4, float(cadangan_bahan)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 4).font = Font(bold=True, color="006100")
            ws.cell(row, 4).border = border_all
            ws.cell(row, 5, "Cadangan").border = border_all
            ws.cell(row, 5).alignment = Alignment(horizontal="center")
            row += 1
        
        # Sewa Dapur Row
        ws.cell(row, 1, "💰 Sewa Dapur").border = border_all
        ws.cell(row, 1).font = Font(bold=True, color="0D47A1")
        ws.cell(row, 4, float(sewa_dapur)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 4).font = Font(bold=True, color="0D47A1")
        ws.cell(row, 4).border = border_all
        ws.cell(row, 5, "Profit Pemilik").border = border_all
        ws.cell(row, 5).alignment = Alignment(horizontal="center")
        row += 1
            
        # Total Row
        total_surplus = total_pagu_all - total_riil_all
        ws.cell(row, 1, "TOTAL EFISIENSI (Pagu - Riil)").font = Font(bold=True)
        ws.cell(row, 1).border = border_all
        ws.cell(row, 2, float(total_pagu_all)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).border = border_all
        ws.cell(row, 3, float(total_riil_all)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 3).border = border_all
        ws.cell(row, 4, float(total_surplus)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 4).font = Font(bold=True, color="006100" if total_surplus >= 0 else "FF0000")
        ws.cell(row, 4).border = border_all
        ws.cell(row, 5, "SURPLUS" if total_surplus >= 0 else "DEFISIT").font = Font(bold=True)
        ws.cell(row, 5).border = border_all
        row += 2

        # Piutang & Hutang
        ws.cell(row, 1, "STATUS PIUTANG & HUTANG").font = Font(name=GLOBAL_FONT, bold=True, color="0D47A1")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1).border = border_all
        ws.cell(row, 2).border = border_all
        row += 1
        
        piutang = float(sum(_d(r.get("amount")) for r in data.get("receivables", [])))
        hutang = float(sum(_d(p.get("amount")) for p in data.get("payables", [])))

        ws.cell(row, 1, "Piutang Cair Bahan Pangan (Aset)"); ws.cell(row, 2, piutang).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 1).font = lbl_font; ws.cell(row, 2).font = val_font
        ws.cell(row, 1).border = border_all; ws.cell(row, 2).border = border_all
        row += 1
        ws.cell(row, 1, "Total Hutang (Payables)");    ws.cell(row, 2, hutang).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 1).font = lbl_font; ws.cell(row, 2).font = val_font
        ws.cell(row, 1).border = border_all; ws.cell(row, 2).border = border_all

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    # ─── Sheet 2: Bulanan (gabungan semua transaksi + penyerahan) ─────────────

    def _fill_bulanan(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (2/10): Filling Bulanan Transactions...")
        from openpyxl.styles import Font, PatternFill, Alignment
        logger.info(
            f"Fill bulanan: {len(data.get('deliveries', []))} deliveries, "
            f"{len(data.get('transactions', []))} transactions, "
            f"{len(data.get('items_all', []))} items"
        )

        bulan_str = BULAN[month]

        # ── Section A: PENYERAHAN MBG (NEW MODEL: Bahan + Ops + Profit) ──
        ws.merge_cells("A1:K1")
        ws["A1"] = f"PENYERAHAN MBG — {bulan_str} {year}"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")

        penyerahan_headers = ["No", "Tanggal", "Sekolah", "Menu", "Porsi",
                              "Bahan (Rp)", "Ops (Rp)", "Sewa Dapur (Rp)", "Total (Rp)"]
        self._write_headers(ws, 2, penyerahan_headers)

        # Build menu lookup
        menu_by_date: Dict[str, str] = {}
        for wm in data.get("weekly_menus", []):
            try:
                ws_date = date.fromisoformat(wm["week_start"])
                d = ws_date + timedelta(days=wm["day_of_week"] - 1)
                menu_by_date[d.isoformat()] = wm.get("menu_name", "—")
            except Exception:
                pass

        row = 3
        total_bahan = total_ops = total_sewa = total_revenue = Decimal("0")
        
        # Get allocation settings for rates
        rate_bahan_sd = data.get("rate_bahan_sd", Decimal("10000"))
        rate_bahan_tk = data.get("rate_bahan_tk", Decimal("8000"))
        rate_ops = data.get("rate_ops_porsi", Decimal("3000"))
        insentif_harian = data.get("insentif_harian", Decimal("6000000"))
        
        seen_dates = set()
        
        dels = data.get("deliveries", [])
        for n, d in enumerate(dels, 1):
            portions = _d(d.get("portions_sent", 0))
            # Safe access for school join (handles dict or list)
            s_data = (d.get("schools") or [{}])[0] if isinstance(d.get("schools"), list) else (d.get("schools") or {})
            school_name = s_data.get("name", "?")
            del_date = d.get("delivery_date", "")
            menu_name = menu_by_date.get(del_date, d.get("menu_name", "—"))
            
            school_level = s_data.get("school_level", "sd_smp")
            rate_bahan = rate_bahan_tk if school_level == "paud_tk" else rate_bahan_sd
            
            bahan = portions * rate_bahan
            ops = portions * rate_ops
            
            # Sewa Dapur: fixed per daily delivery date
            if del_date not in seen_dates:
                sewa = insentif_harian
                seen_dates.add(del_date)
            else:
                sewa = Decimal("0")
            
            total = bahan + ops + sewa

            total_bahan += bahan
            total_ops += ops
            total_sewa += sewa
            total_revenue += total

            ws.cell(row, 1, n)
            ws.cell(row, 2, del_date)
            ws.cell(row, 3, school_name)
            ws.cell(row, 4, menu_name)
            ws.cell(row, 5, portions)
            ws.cell(row, 6, float(bahan)).number_format = '"Rp "#,##0'
            ws.cell(row, 7, float(ops)).number_format = '"Rp "#,##0'
            ws.cell(row, 8, float(sewa)).number_format = '"Rp "#,##0'
            ws.cell(row, 9, float(total)).number_format = '"Rp "#,##0'
            self._alt_row(ws, row, 2)
            row += 1

        if not data.get("deliveries"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada penyerahan MBG bulan ini")
            row += 1

        # Total penyerahan
        self._write_total_row(ws, row, {
            5: sum(d.get("portions_sent", 0) for d in data.get("deliveries", [])),
            6: float(total_bahan), 
            7: float(total_ops), 
            8: float(total_sewa),
            9: float(total_revenue)
        }, currency_cols=[6, 7, 8, 9])
        row += 2

        # ── Section B: PEMBELIAN / PENGELUARAN ──
        section_b_start = row
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = f"PEMBELIAN & PENGELUARAN — {bulan_str} {year}"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        pembelian_headers = ["No", "Tanggal", "Supplier", "Item", "Qty", "Satuan",
                             "Harga Satuan", "Subtotal",
                             "Total Bayar", "Metode Bayar", "Kategori Juknis", "Foto Nota"]
        self._write_headers(ws, row, pembelian_headers)
        row += 1

        # Index items by transaction_id
        items_by_trx: Dict[str, List[dict]] = {}
        for item in data.get("items_all", []):
            tid = item.get("transaction_id", "")
            items_by_trx.setdefault(tid, []).append(item)

        n = 0
        tot_sub = tot_pph = tot_ppn = tot_pay = Decimal("0")

        for trx in data.get("transactions", []):
            tid = trx["id"]
            del_date = trx.get("date", "")
            sup_data = trx.get("suppliers") or {}
            supplier = sup_data.get("name") if isinstance(sup_data, dict) else None
            supplier = supplier or trx.get("nama_toko") or "Tanpa Supplier"
            method = trx.get("payment_method", "tunai") or "tunai"
            sup_id = trx.get("supplier_id")
            is_pkp = data["supplier_pkp"].get(sup_id, False) if sup_id else False

            trx_items = items_by_trx.get(tid, [])

            if not trx_items:
                # Transaksi tanpa item — tampilkan 1 baris
                n += 1
                total_trx = float(_d(trx.get("total")))
                ws.cell(row, 1, n)
                ws.cell(row, 2, del_date)
                ws.cell(row, 3, supplier)
                ws.cell(row, 4, "(tanpa detail item)")
                ws.cell(row, 5, "")
                ws.cell(row, 6, "")
                ws.cell(row, 7, "")
                ws.cell(row, 8, total_trx).number_format = '"Rp "#,##0'
                ws.cell(row, 9, total_trx).number_format = '"Rp "#,##0'
                ws.cell(row, 10, method)
                jcat = trx.get("juknis_category", "")
                jcat_label = {"bahan_pangan": "Bahan Pangan", "operasional": "Operasional", "kemasan": "Kemasan", "alat_tulis": "Alat Tulis"}.get(jcat, jcat or "Lainnya")
                ws.cell(row, 11, jcat_label)
                photo_url = trx.get("photo_url")
                if photo_url:
                    from openpyxl.styles import Font as _F
                    cell = ws.cell(row, 12, "Lihat Foto")
                    cell.hyperlink = photo_url
                    cell.font = _F(color="0563C1", underline="single")
                else:
                    ws.cell(row, 12, "")
                tot_sub += _d(trx.get("total"))
                tot_pay += _d(trx.get("total"))
                self._alt_row(ws, row, 2)
                row += 1
                continue

            for item in trx_items:
                n += 1
                nama = item.get("product_name", "?")
                qty  = float(_d(item.get("qty")))
                sat  = item.get("unit", "pcs")
                harga = float(_d(item.get("price")))
                subtotal = _d(item.get("subtotal")) or (_d(item.get("qty")) * _d(item.get("price")))
                bayar = subtotal

                tot_sub += subtotal
                tot_pay += bayar

                ws.cell(row, 1, n)
                ws.cell(row, 2, del_date)
                ws.cell(row, 3, supplier)
                ws.cell(row, 4, nama)
                ws.cell(row, 5, qty)
                ws.cell(row, 6, sat)
                ws.cell(row, 7, harga).number_format = '"Rp "#,##0'
                ws.cell(row, 8, float(subtotal)).number_format = '"Rp "#,##0'
                ws.cell(row, 9, float(bayar)).number_format = '"Rp "#,##0'
                ws.cell(row, 10, method)
                jcat = trx.get("juknis_category", "")
                jcat_label = {"bahan_pangan": "Bahan Pangan", "operasional": "Operasional", "kemasan": "Kemasan", "alat_tulis": "Alat Tulis"}.get(jcat, jcat or "Lainnya")
                ws.cell(row, 11, jcat_label)
                photo_url = trx.get("photo_url")
                if photo_url:
                    from openpyxl.styles import Font as _F
                    cell = ws.cell(row, 12, "Lihat Foto")
                    cell.hyperlink = photo_url
                    cell.font = _F(color="0563C1", underline="single")
                else:
                    ws.cell(row, 12, "")
                self._alt_row(ws, row, 2)
                row += 1

        if not data.get("transactions"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada transaksi bulan ini")
            row += 1

        self._write_total_row(ws, row, {
            8: float(tot_sub), 9: float(tot_pay),
        }, max_col=12)

        self._set_column_widths(ws)
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = "A2:L2"

    # ─── Sheet 3: Stok ────────────────────────────────────────────────────────

    def _fill_stok(self, ws, data: dict) -> None:
        logger.info("Sheet (3/10): Filling Stok Bahan...")
        logger.info(f"Fill stok: {len(data.get('products', []))} products, {len(data.get('stock_history', []))} history records")
        headers = ["No", "Nama Bahan", "Kategori", "Satuan",
                   "Harga", "Stok Awal Bulan", "Masuk", "Keluar",
                   "Stok Akhir", "Nilai Stok"]
        self._write_headers(ws, 1, headers)

        # Build net change per product
        in_by_prod:  Dict[str, Decimal] = {}
        out_by_prod: Dict[str, Decimal] = {}
        for h in data.get("stock_history", []):
            pid = h.get("product_id", "")
            delta = _d(h.get("change_qty"))
            if delta >= 0:
                in_by_prod[pid]  = in_by_prod.get(pid, Decimal("0"))  + delta
            else:
                out_by_prod[pid] = out_by_prod.get(pid, Decimal("0")) + abs(delta)

        row = 2
        
        # Filter out 'produk_jadi' and 'komponen' so only raw materials and packaging are listed
        valid_products = [p for p in data.get("products", []) if p.get("category") not in ("produk_jadi", "komponen")]
        
        for n, p in enumerate(valid_products, 1):
            pid = p.get("id", "")
            harga = _d(p.get("harga"))
            stk_akhir = _d(p.get("stock_qty"))
            masuk  = in_by_prod.get(pid, Decimal("0"))
            keluar = out_by_prod.get(pid, Decimal("0"))
            stk_awal = stk_akhir - masuk + keluar
            factor = float(p.get("conversion_factor") or 1) or 1
            
            # Nilai stok is based on display unit qty * harga
            stk_akhir_display = float(stk_akhir) / factor
            nilai = stk_akhir_display * float(harga)

            ws.cell(row, 1, n)
            ws.cell(row, 2, p.get("name"))
            ws.cell(row, 3, p.get("category"))
            ws.cell(row, 4, p.get("display_unit") or p.get("unit"))
            ws.cell(row, 5, float(harga)).number_format = '"Rp "#,##0'
            ws.cell(row, 6, round(float(stk_awal) / factor, 3))
            ws.cell(row, 7, round(float(masuk) / factor, 3))
            ws.cell(row, 8, round(float(keluar) / factor, 3))
            ws.cell(row, 9, round(stk_akhir_display, 3))
            ws.cell(row, 10, round(nilai, 2)).number_format = ACCOUNTING_FORMAT
            self._alt_row(ws, row, 10)
            row += 1

        if not data.get("products"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada data produk/bahan")
            row += 1

        self._set_column_widths(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{row-1}"

    # ─── Sheet 4: Laporan Pemerintah ──────────────────────────────────────────

    def _fill_laporan_pemerintah(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (4/10): Filling Laporan Pemerintah (Advanced Financials)...")
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        bulan_str = BULAN[month]
        tenant_name = data["tenant"].get("name", "SPPG")

        # Styles
        border_all = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.merge_cells("A1:E1")
        ws["A1"] = f"LAPORAN KEUANGAN KONSOLIDASI (PEMERINTAH) — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "A. IDENTITAS PROGRAM"
        ws["A3"].font = Font(bold=True, size=11)
        ws["A4"] = "Nama SPPG"; ws["B4"] = tenant_name
        ws["A5"] = "Periode"; ws["B5"] = f"{bulan_str} {year}"

        # Get settings and data
        rate_bahan_sd = data.get("rate_bahan_sd", Decimal("10000"))
        rate_bahan_tk = data.get("rate_bahan_tk", Decimal("8000"))
        rate_ops = data.get("rate_ops_porsi", Decimal("3000"))
        insentif_harian = data.get("insentif_harian", Decimal("6000000"))
        working_days = data.get("working_days_month", 26)

        total_porsi = _d(data.get("total_portions_all", 0))
        pagu_ops = total_porsi * rate_ops
        pagu_sewa = insentif_harian * working_days

        pagu_bahan = Decimal("0")
        for d in data.get("deliveries", []):
            p = _d(d.get("portions_sent", 0))
            s_data = (d.get("schools") or [{}])[0] if isinstance(d.get("schools"), list) else (d.get("schools") or {})
            sl = s_data.get("school_level", "sd_smp")
            pagu_bahan += p * (rate_bahan_tk if sl == "paud_tk" else rate_bahan_sd)

        total_pagu = pagu_bahan + pagu_ops + pagu_sewa

        # Realisasi
        riil_bahan = sum(_d(t.get("total")) for t in data.get("transactions", []) 
                        if str(t.get("juknis_category") or "").lower() == "bahan_pangan")
        riil_ops_base = sum(_d(cost.get("amount")) for cost in data.get("operational_costs", []))
        riil_gaji = sum(_d(item.get("net_amount")) for item in data.get("payroll_items", []))
        riil_ops_trx = sum(_d(t.get("total")) for t in data.get("transactions", []) 
                         if str(t.get("juknis_category") or "").lower() != "bahan_pangan")
        riil_ops = riil_ops_base + riil_gaji + riil_ops_trx
        
        # Profit / Sewa Dapur is fixed
        riil_sewa = pagu_sewa
        
        total_riil = riil_bahan + riil_ops + riil_sewa
        sisa_dana = total_pagu - total_riil

        ws["A7"] = "B. VOLUME PRODUKSI"
        ws["A7"].font = Font(bold=True, size=11)
        ws["A8"] = "Total Porsi Terkirim"; ws["B8"] = total_porsi
        ws["A9"] = "Hari Kerja Aktif"; ws["B9"] = working_days

        row = 11
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "C. RINCIAN PENGGUNAAN DANA (ANGGARAN vs REALISASI)"
        ws[f"A{row}"].font = Font(bold=True, size=11)
        row += 1

        headers = ["Komponen Biaya", "Anggaran (Pagu)", "Realisasi (Riil)", "Varian (Selisih)", "Status"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row, i)
            cell.value = h
            cell.font = Font(bold=True)
            cell.border = border_all
        row += 1

        rows_data = [
            ("1. Bahan Baku Pangan", pagu_bahan, riil_bahan),
            ("2. Biaya Operasional & Tenaga Kerja (Rp 3k/porsi)", pagu_ops, riil_ops),
            ("3. Sewa Dapur / Keuntungan Pengelolaan", pagu_sewa, riil_sewa),
        ]

        for name, pagu, riil in rows_data:
            diff = pagu - riil
            status = "✅ Sesuai/Hemat" if diff >= 0 else "⚠️ Over"
            if name.startswith("3."):
                status = "Fixed Profit"
            ws.cell(row, 1, name).border = border_all
            ws.cell(row, 2, float(pagu)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 2).border = border_all
            ws.cell(row, 3, float(riil)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 3).border = border_all
            ws.cell(row, 4, float(diff)).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 4).border = border_all
            ws.cell(row, 5, status).border = border_all
            if diff < 0:
                ws.cell(row, 4).font = Font(color="FF0000", bold=True)
                ws.cell(row, 5).font = Font(color="FF0000")
            elif diff > 0 and not name.startswith("3."):
                ws.cell(row, 4).font = Font(color="006100", bold=True)
                ws.cell(row, 5).font = Font(color="006100")
            row += 1

        row += 1
        ws.cell(row, 1, "TOTAL PENGGUNAAN DANA").font = Font(bold=True)
        ws.cell(row, 1).border = border_all
        ws.cell(row, 2, float(total_pagu)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).border = border_all
        ws.cell(row, 3, float(total_riil)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 3).border = border_all
        ws.cell(row, 4, float(sisa_dana)).number_format = ACCOUNTING_FORMAT
        ws.cell(row, 4).font = Font(bold=True, color="006100" if sisa_dana >= 0 else "FF0000")
        ws.cell(row, 4).border = border_all
        ws.cell(row, 5, "SURPLUS" if sisa_dana >= 0 else "DEFISIT").font = Font(bold=True)
        ws.cell(row, 5).border = border_all

        row += 3
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "D. DETAIL BIAYA OPERASIONAL & TENAGA KERJA"
        ws[f"A{row}"].font = Font(bold=True, size=11)
        row += 1
        
        ws.cell(row, 1, "Kategori Biaya Operasional").font = Font(bold=True)
        ws.cell(row, 3, "Nominal Riil").font = Font(bold=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="right")
        row += 1
        
        ws.cell(row, 1, "1. Gaji & Upah Karyawan")
        ws.cell(row, 3, float(riil_gaji)).number_format = ACCOUNTING_FORMAT
        row += 1
        ws.cell(row, 1, "2. Pembelian Operasional (BBM, Gas, Listrik)")
        ws.cell(row, 3, float(riil_ops_base)).number_format = ACCOUNTING_FORMAT
        row += 1
        ws.cell(row, 1, "3. Pembelian Non-Bahan Lainnya (Nota)")
        ws.cell(row, 3, float(riil_ops_trx)).number_format = ACCOUNTING_FORMAT
        row += 1

        self._set_column_widths(ws)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 20

    # ─── Sheet 5: Riwayat Nota ────────────────────────────────────────────────

    def _fill_riwayat_nota(self, ws, data: dict) -> None:
        logger.info("Sheet (5/10): Filling Riwayat Nota (Detail Item)...")
        logger.info(f"Fill riwayat nota: {len(data.get('transactions', []))} transactions, {len(data.get('items_all', []))} items")
        headers = ["No", "Tanggal", "Supplier", "Item", "Qty", "Satuan",
                   "Harga", "Subtotal", "Total Nota", "Metode Bayar"]
        self._write_headers(ws, 1, headers)

        items_by_trx: Dict[str, List[dict]] = {}
        for item in data.get("items_all", []):
            tid = item.get("transaction_id", "")
            items_by_trx.setdefault(tid, []).append(item)

        row = 2
        n = 0
        for trx in data.get("transactions", []):
            tid = trx["id"]
            del_date = trx.get("date", "")
            sup_data = trx.get("suppliers") or {}
            supplier = sup_data.get("name") if isinstance(sup_data, dict) else None
            supplier = supplier or trx.get("nama_toko") or "Tanpa Supplier"
            method = trx.get("payment_method", "tunai") or "tunai"
            total_nota = float(_d(trx.get("total")))

            trx_items = items_by_trx.get(tid, [])

            if not trx_items:
                n += 1
                ws.cell(row, 1, n)
                ws.cell(row, 2, del_date)
                ws.cell(row, 3, supplier)
                ws.cell(row, 4, "(tanpa detail)")
                ws.cell(row, 9, total_nota).number_format = '"Rp "#,##0'
                ws.cell(row, 10, method)
                self._alt_row(ws, row, 10)
                row += 1
                continue

            for item in trx_items:
                n += 1
                nama = item.get("product_name", "?")
                qty = float(_d(item.get("qty")))
                sat = item.get("unit", "pcs")
                harga = float(_d(item.get("price")))
                subtotal = float(_d(item.get("subtotal")) or (_d(item.get("qty")) * _d(item.get("price"))))

                ws.cell(row, 1, n)
                ws.cell(row, 2, del_date)
                ws.cell(row, 3, supplier)
                ws.cell(row, 4, nama)
                ws.cell(row, 5, qty)
                ws.cell(row, 6, sat)
                ws.cell(row, 7, harga).number_format = '"Rp "#,##0'
                ws.cell(row, 8, subtotal).number_format = '"Rp "#,##0'
                ws.cell(row, 9, total_nota).number_format = '"Rp "#,##0'
                ws.cell(row, 10, method)
                self._alt_row(ws, row, 10)
                row += 1

        if not data.get("transactions"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada riwayat nota bulan ini")
            row += 1

        self._set_column_widths(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{row-1}"

    # ─── Sheet 6: Operasional & Upah ──────────────────────────────────────────

    def _fill_ops_upah(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (6/10): Filling Operasional & Upah (Analisis Pagu vs Riil)...")
        from openpyxl.styles import Font, PatternFill, Alignment
        bulan_str = BULAN[month]

        # ── Section A: DETAIL ANALISIS TARGET (PAGU) vs REALISASI (RIIL) ──
        ws.merge_cells("A1:E1")
        ws["A1"] = f"DETAIL ANALISIS OPERASIONAL & UPAH — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")

        sum_headers = ["Kategori", "Anggaran (Pagu Juknis)", "Realisasi (Riil)", "Efisiensi (Tabungan SPPG)", "Status"]
        self._write_headers(ws, 2, sum_headers)

        # Target Operasional + Gaji = total_porsi × Rp 3.000 (satu sumber dana)
        total_portions = _d(data.get("total_portions_all", 0))
        rate_ops = data.get("rate_ops_porsi", Decimal("3000"))
        target_ops_gaji = total_portions * rate_ops  # Rp 3k/porsi covers ALL ops + gaji

        # Realisasi: Ops Riil + Gaji + Nota Non-Bahan
        actual_ops_base = sum(_d(cost.get("amount")) for cost in data.get("operational_costs", []))
        actual_gaji = sum(_d(item.get("net_amount")) for item in data.get("payroll_items", []))
        ops_from_trx = sum(_d(t.get("total")) for t in data.get("transactions", []) 
                          if str(t.get("juknis_category") or "").lower() != "bahan_pangan")
        actual_ops_gaji = actual_ops_base + actual_gaji + ops_from_trx

        # Sewa Dapur (info only — keuntungan pemilik, bukan pengeluaran)
        insentif_harian = data.get("insentif_harian", Decimal("6000000"))
        working_days = data.get("working_days_month", 26)
        sewa_dapur = insentif_harian * working_days

        row = 3
        # Single combined row: Ops + Gaji from Rp 3k/porsi
        selisih = target_ops_gaji - actual_ops_gaji
        status = "✅ Hemat" if selisih >= 0 else "⚠️ Over"
        ws.cell(row, 1, f"1. Operasional + Gaji (Rp {int(rate_ops):,}/porsi × {int(total_portions):,} porsi)")
        ws.cell(row, 2, float(target_ops_gaji)).number_format = '"Rp "#,##0'
        ws.cell(row, 3, float(actual_ops_gaji)).number_format = '"Rp "#,##0'
        ws.cell(row, 4, float(selisih)).number_format = '"Rp "#,##0'
        ws.cell(row, 5, status)
        if selisih < 0:
            ws.cell(row, 4).font = Font(color="FF0000", bold=True)
            ws.cell(row, 5).font = Font(color="FF0000")
        else:
            ws.cell(row, 4).font = Font(color="006400", bold=True)
            ws.cell(row, 5).font = Font(color="006400")
        row += 1
        
        # Detail breakdown within Rp 3k/porsi
        ws.cell(row, 1, "   ├ Biaya Operasional (BBM, Gas, Logistik)")
        ws.cell(row, 3, float(actual_ops_base)).number_format = '"Rp "#,##0'
        ws.cell(row, 3).font = Font(italic=True, color="555555")
        row += 1
        ws.cell(row, 1, "   ├ Nota Non-Bahan")
        ws.cell(row, 3, float(ops_from_trx)).number_format = '"Rp "#,##0'
        ws.cell(row, 3).font = Font(italic=True, color="555555")
        row += 1
        ws.cell(row, 1, "   └ Gaji Karyawan")
        ws.cell(row, 3, float(actual_gaji)).number_format = '"Rp "#,##0'
        ws.cell(row, 3).font = Font(italic=True, color="555555")
        row += 1
        
        # Sewa Dapur info row
        row += 1
        ws.cell(row, 1, "💰 Sewa Dapur")
        ws.cell(row, 1).font = Font(bold=True, color="0D47A1")
        ws.cell(row, 2, float(sewa_dapur)).number_format = '"Rp "#,##0'
        ws.cell(row, 2).font = Font(bold=True, color="0D47A1")
        ws.cell(row, 5, "Keuntungan Pemilik")
        ws.cell(row, 5).font = Font(bold=True, color="0D47A1")

        row += 1

        # ── Section B: DETAIL BIAYA OPERASIONAL ──
        ops_start = row
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = f"DETAIL BIAYA OPERASIONAL — {bulan_str} {year}"
        ws[f"A{row}"].font = Font(name=GLOBAL_FONT, bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        ops_headers = ["No", "Tanggal", "Nama Biaya", "Kategori", "Nominal", "Rutin?", "Catatan"]
        self._write_headers(ws, row, ops_headers)
        row += 1

        total_ops_detail = Decimal("0")
        n = 0
        
        for cost in data.get("operational_costs", []):
            n += 1
            amt = float(_d(cost.get("amount")))
            total_ops_detail += _d(cost.get("amount"))
            
            ws.cell(row, 1, n)
            ws.cell(row, 2, cost.get("cost_date", ""))
            ws.cell(row, 3, cost.get("name", ""))
            ws.cell(row, 4, cost.get("category", ""))
            ws.cell(row, 5, amt).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 6, "Ya" if cost.get("is_recurring") else "Tidak")
            ws.cell(row, 7, cost.get("notes", ""))
            self._alt_row(ws, row, 7)
            row += 1

        if not data.get("operational_costs"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada biaya operasional bulan ini")
            row += 1

        self._write_total_row(ws, row, {5: float(total_ops_detail)}, max_col=7)
        row += 2

        # ── Section C: GAJI & UPAH ──
        upah_start = row
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = f"GAJI & UPAH (BERDASARKAN PERIODE) — {bulan_str} {year}"
        ws[f"A{row}"].font = Font(name=GLOBAL_FONT, bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        upah_headers = ["No", "Periode", "Status", "Mulai", "Nama", "Jabatan", "Hadir", "Alpa", "Potongan", "Total Terima"]
        self._write_headers(ws, row, upah_headers)
        row += 1

        total_upah_detail = Decimal("0")
        n = 0
        
        for item in data.get("payroll_items", []):
            n += 1
            period = item.get("payroll_periods", {})
            emp = item.get("employees", {})
            net = float(_d(item.get("net_amount")))
            total_upah_detail += _d(item.get("net_amount"))
            
            ws.cell(row, 1, n)
            ws.cell(row, 2, period.get("name", ""))
            ws.cell(row, 3, period.get("status", ""))
            ws.cell(row, 4, period.get("start_date", ""))
            ws.cell(row, 5, emp.get("name", ""))
            ws.cell(row, 6, item.get("position_name", ""))
            ws.cell(row, 7, item.get("present_days", 0))
            ws.cell(row, 8, item.get("absent_days", 0))
            ws.cell(row, 9, float(_d(item.get("deductions")))).number_format = ACCOUNTING_FORMAT
            ws.cell(row, 10, net).number_format = ACCOUNTING_FORMAT
            self._alt_row(ws, row, 10)
            row += 1
            
        if not data.get("payroll_items"):
            ws.cell(row, 1, "—")
            ws.cell(row, 2, "Tidak ada data gaji bulan ini")
            row += 1
            
        self._write_total_row(ws, row, {10: float(total_upah_detail)}, max_col=10)
        
        self._set_column_widths(ws)
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A{ops_start+1}:G{ops_start+1}" # Just a filter on one of the headers

    # ─── Helpers ──────────────────────────────────────────────────────────────

    def _write_headers(self, ws, row: int, headers: list) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        font = Font(name=GLOBAL_FONT, bold=True, color="FFFFFF", size=GLOBAL_FONT_SIZE)
        border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        for i, h in enumerate(headers, 1):
            c = ws.cell(row, i, h)
            c.font = font
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    def _apply_header_style(self, ws, row: int) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        fill = PatternFill("solid", fgColor=HDR_FILL_COLOR)
        font = Font(name=GLOBAL_FONT, bold=True, color="FFFFFF", size=GLOBAL_FONT_SIZE)
        border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        for cell in ws[row]:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def _alt_row(self, ws, row: int, max_col: int) -> None:
        from openpyxl.styles import PatternFill, Font, Border, Side
        fill = None
        if row % 2 == 0:
            fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
        
        border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        font = Font(name=GLOBAL_FONT, size=GLOBAL_FONT_SIZE)
        
        for c in range(1, max_col + 1):
            cell = ws.cell(row, c)
            if fill:
                cell.fill = fill
            cell.border = border
            if not cell.font or cell.font.name != GLOBAL_FONT:
                cell.font = font

    def _write_total_row(self, ws, row: int, col_values: dict, max_col: int = 0, currency_cols: Optional[List[int]] = None) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        fill = PatternFill("solid", fgColor=TOTAL_FILL_COLOR)
        font = Font(name=GLOBAL_FONT, bold=True, size=GLOBAL_FONT_SIZE)
        border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )

        # Apply style to the whole row up to max_col
        limit = max_col if max_col > 0 else max(col_values.keys()) if col_values else 10
        for c in range(1, limit + 1):
            cell = ws.cell(row, c)
            cell.fill = fill
            cell.font = font
            cell.border = border

        ws.cell(row, 1, "TOTAL")
        for col, val in col_values.items():
            if val is not None:
                c = ws.cell(row, col, val)
                if isinstance(val, (int, float, Decimal)):
                    if currency_cols is None or col in currency_cols:
                        c.number_format = ACCOUNTING_FORMAT
                    else:
                        c.number_format = '#,##0'

    def _set_column_widths(self, ws, padding: int = 5) -> None:
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Skip long text in title rows or merged rows to prevent giant column widths
                        if len(val_str) > 25 and cell.row in [1, 2, 3]:
                            continue
                        
                        fmt = str(cell.number_format or "")
                        # If cell contains formula, it will evaluate to a value, we must give it space
                        is_formula = val_str.startswith("=")
                        is_numeric = isinstance(cell.value, (int, float, Decimal)) or is_formula
                        
                        extra = 10 if ("Rp" in fmt or "#" in fmt or "0" in fmt) and is_numeric else 0
                        if is_formula:
                            # Formula length might be short but output might be long (e.g. Rp 1.500.000)
                            length = 16
                        else:
                            length = len(val_str) + extra
                        
                        max_len = max(max_len, length)
                except Exception:
                    pass
            # Default minimum 12, max 50
            ws.column_dimensions[col_letter].width = min(max(max_len + padding, 12), 50)

    # ─── Storage + DB ─────────────────────────────────────────────────────────

    def _upload_to_storage(
        self, supabase, tenant_id: str, year: int, month: int, content: bytes
    ) -> Optional[str]:
        file_name = f"{year}-{month:02d}_pembukuan.xlsx"
        storage_path = f"{tenant_id}/excel/{file_name}"
        bucket = "nota-photos"

        try:
            # Use upsert: true to overwrite if exists
            supabase.storage.from_(bucket).upload(
                path=storage_path,
                file=content,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "upsert": "true"
                },
            )
            url = supabase.storage.from_(bucket).get_public_url(storage_path)
            return url if isinstance(url, str) else str(url)
        except Exception as e:
            logger.error(f"Upload Excel gagal: {e}")
            return None

    def _upsert_excel_files(
        self, supabase, tenant_id: str, period: str,
        status: str = "generating",
        file_url: Optional[str] = None,
        error_message: Optional[str] = None,
    ) -> None:
        now = datetime.now(timezone.utc).isoformat()
        try:
            year, month = map(int, period.split("-"))
        except:
            year, month = 0, 0
            
        row: Dict[str, Any] = {
            "tenant_id": tenant_id,
            "period": period,
            "year": year,
            "month": month,
            "status": status,
            "last_updated": now,
            "file_url": file_url or "",
        }
        if error_message:
            row["error_message"] = error_message

        try:
            supabase.table("excel_files").upsert(
                row, on_conflict="tenant_id,period"
            ).execute()
        except Exception as e:
            logger.warning(f"upsert excel_files gagal: {e}")

    # ═══════════════════════════════════════════════════════════════════════
    # COMPLIANCE SHEETS — Modul 21.5e
    # ═══════════════════════════════════════════════════════════════════════

    def _fill_hygiene(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (11/15): Filling Checklist Higiene...")
        """Sheet 11: Checklist Higiene Bulanan."""
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]

        ws.merge_cells("A1:J1")
        ws["A1"] = f"LAPORAN CHECKLIST HIGIENE — {bulan_str} {year}"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Tanggal", "Lantai/Dinding", "Suhu Penyimpanan", "Peralatan Masak",
            "Atribut Personel", "Air Bersih", "Area Penyimpanan", "Tempat Sampah",
            "Status", "Catatan"
        ]
        hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        checks = data.get("hygiene_checks", [])
        area_map = {
            "Lantai": 2, "Suhu": 3, "Peralatan": 4, "Personel": 5,
            "Air": 6, "Penyimpanan": 7, "Sampah": 8,
        }

        for idx, check in enumerate(checks, start=4):
            ws.cell(row=idx, column=1, value=check.get("check_date", ""))

            items = check.get("items", [])
            for item in items:
                area_name = item.get("area", "")
                for key, col in area_map.items():
                    if key.lower() in area_name.lower():
                        icon = "✓" if item.get("status") == "baik" else "⚠️"
                        ws.cell(row=idx, column=col, value=icon)
                        break

            status = check.get("overall_status", "")
            status_cell = ws.cell(row=idx, column=9, value=status.upper())
            
            # Dashboard Biru Modern Style: Green for Safe, Red for Incident
            if status == "layak":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(name=GLOBAL_FONT, color="006100", bold=True)
            elif status == "tidak_layak":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(name=GLOBAL_FONT, color="9C0006", bold=True)
            else: # perlu_perbaikan
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(name=GLOBAL_FONT, color="9C6500", bold=True)

            ws.cell(row=idx, column=10, value=check.get("notes", ""))
            self._alt_row(ws, idx, 10)

        ws.freeze_panes = "A4"
        self._set_column_widths(ws)

    def _fill_suhu(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (12/15): Filling Monitoring Suhu...")
        """Sheet 12: Monitoring Suhu."""
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]

        ws.merge_cells("A1:F1")
        ws["A1"] = f"LAPORAN MONITORING SUHU — {bulan_str} {year}"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Tanggal", "Jam", "Area", "Suhu (°C)", "Status", "Catatan"]
        hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        logs = data.get("temperature_logs", [])
        for idx, log in enumerate(logs, start=4):
            ws.cell(row=idx, column=1, value=log.get("log_date", ""))
            ws.cell(row=idx, column=2, value=log.get("log_time", ""))
            ws.cell(row=idx, column=3, value=log.get("area", ""))

            temp_cell = ws.cell(row=idx, column=4, value=log.get("temperature"))
            temp_cell.number_format = '0.0'

            status_cell = ws.cell(row=idx, column=5)
            if not log.get("is_normal"):
                status_cell.value = "⚠️ ABNORMAL"
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(name=GLOBAL_FONT, color="9C0006", bold=True)
            else:
                status_cell.value = "✓ Normal"
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(name=GLOBAL_FONT, color="006100", bold=True)

            ws.cell(row=idx, column=6, value=log.get("notes", ""))
            self._alt_row(ws, idx, 6)

        ws.freeze_panes = "A4"
        self._set_column_widths(ws)

    def _fill_bank_sampel(self, ws, data: dict) -> None:
        logger.info("Sheet (13/15): Filling Bank Sampel...")
        """Sheet 13: Bank Sampel Makanan."""
        from openpyxl.styles import Font, Alignment, PatternFill

        ws.merge_cells("A1:H1")
        ws["A1"] = "DAFTAR BANK SAMPEL MAKANAN"
        ws["A1"].font = Font(name=GLOBAL_FONT, bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Kode Sampel", "Menu", "Tanggal", "Jam Ambil", "Berat (g)", "Kadaluarsa", "Status", "Sisa Waktu"]
        hdr_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        samples = data.get("food_samples", [])
        now = datetime.now(timezone.utc)

        for idx, sample in enumerate(samples, start=4):
            ws.cell(row=idx, column=1, value=sample.get("sample_code", ""))
            ws.cell(row=idx, column=2, value=sample.get("menu_name", ""))
            ws.cell(row=idx, column=3, value=sample.get("sample_date", ""))
            ws.cell(row=idx, column=4, value=sample.get("taken_at", ""))
            ws.cell(row=idx, column=5, value=sample.get("weight_gram"))
            ws.cell(row=idx, column=6, value=sample.get("expires_at", ""))

            status_cell = ws.cell(row=idx, column=7)
            hours_cell = ws.cell(row=idx, column=8)

            try:
                exp_str = sample.get("expires_at", "")
                if exp_str:
                    if "+" not in exp_str and "Z" not in exp_str:
                        exp_str += "+00:00"
                    expires_dt = datetime.fromisoformat(exp_str.replace("Z", "+00:00"))
                    hours_left = (expires_dt - now).total_seconds() / 3600

                    if hours_left < 0:
                        status_cell.value = "❌ KADALUARSA"
                        status_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                        status_cell.font = Font(color="FFFFFF")
                        hours_cell.value = "Expired"
                    elif hours_left <= 4:
                        status_cell.value = "⚠️ Segera"
                        status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                        hours_cell.value = f"{hours_left:.1f} jam"
                    else:
                        status_cell.value = "✓ Aman"
                        status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                        hours_cell.value = f"{hours_left:.1f} jam"
                else:
                    status_cell.value = "—"
                    hours_cell.value = "—"
            except Exception:
                status_cell.value = sample.get("status", "")
                hours_cell.value = "—"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        for col_letter in ["C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 14

    def _fill_sisa_makanan(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (14/15): Filling Laporan Sisa Makanan...")
        """Sheet 14: Laporan Sisa Makanan."""
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]

        ws.merge_cells("A1:H1")
        ws["A1"] = f"LAPORAN SISA MAKANAN — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Tanggal", "Sekolah", "Porsi Kirim", "Porsi Konsumsi", "% Sisa", "Skor Comstock", "Alasan", "Catatan"]
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        reports = data.get("food_waste_reports", [])

        total_sent = 0
        total_consumed = 0

        for idx, rep in enumerate(reports, start=4):
            ws.cell(row=idx, column=1, value=rep.get("report_date", ""))
            school_info = rep.get("schools") or {}
            ws.cell(row=idx, column=2, value=school_info.get("name", "") if isinstance(school_info, dict) else "")
            ws.cell(row=idx, column=3, value=rep.get("portions_sent", 0))
            ws.cell(row=idx, column=4, value=rep.get("portions_consumed", 0))

            pct_cell = ws.cell(row=idx, column=5, value=rep.get("waste_pct", 0))
            pct_cell.number_format = '0.0"%"'

            score = rep.get("comstock_score", 0)
            ws.cell(row=idx, column=6, value=f"⭐{score}" if score else "")
            ws.cell(row=idx, column=7, value=rep.get("waste_reason", ""))
            ws.cell(row=idx, column=8, value=rep.get("notes", ""))

            total_sent += rep.get("portions_sent", 0)
            total_consumed += rep.get("portions_consumed", 0)

        # Summary row
        if reports:
            summary_row = len(reports) + 5
            ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=total_sent).font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=total_consumed).font = Font(bold=True)
            total_pct = round((total_sent - total_consumed) / total_sent * 100, 1) if total_sent > 0 else 0
            pct_summary = ws.cell(row=summary_row, column=5, value=total_pct)
            pct_summary.number_format = '0.0"%"'
            pct_summary.font = Font(bold=True)
            pct_summary.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 15

    def _fill_insiden(self, ws, data: dict, year: int, month: int) -> None:
        logger.info("Sheet (15/15): Filling Laporan Insiden...")
        """Sheet 15: Laporan Insiden KLB."""
        from openpyxl.styles import Font, Alignment, PatternFill
        bulan_str = BULAN[month]

        ws.merge_cells("A1:J1")
        ws["A1"] = f"LAPORAN INSIDEN KLB/KERACUNAN — {bulan_str} {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Kode Insiden", "Tanggal", "Sekolah", "Lokasi", "Korban", "Gejala", "Tindakan", "Sampel Diamankan", "Status", "Hasil Investigasi"]
        hdr_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        incidents = data.get("incident_reports", [])

        for idx, inc in enumerate(incidents, start=4):
            ws.cell(row=idx, column=1, value=inc.get("incident_code", ""))
            ws.cell(row=idx, column=2, value=(inc.get("incident_time", "") or "")[:10])
            school_info = inc.get("schools") or {}
            ws.cell(row=idx, column=3, value=school_info.get("name", "") if isinstance(school_info, dict) else "")
            ws.cell(row=idx, column=4, value=inc.get("location", ""))
            ws.cell(row=idx, column=5, value=inc.get("victim_count", 0))
            ws.cell(row=idx, column=6, value=", ".join(inc.get("symptoms", []) or []))
            ws.cell(row=idx, column=7, value=inc.get("first_action", ""))

            sample_cell = ws.cell(row=idx, column=8)
            sample_cell.value = "✓ Ya" if inc.get("sample_secured") else "✗ Tidak"
            if inc.get("sample_secured"):
                sample_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            status_cell = ws.cell(row=idx, column=9, value=(inc.get("status", "") or "").upper())
            status_val = inc.get("status", "")
            if status_val == "investigasi":
                status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif status_val == "selesai":
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif status_val == "ditutup":
                status_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            ws.cell(row=idx, column=10, value=inc.get("investigation_result", ""))

        if not incidents:
            ws.cell(row=4, column=1, value="Tidak ada insiden pada bulan ini")
            ws.merge_cells("A4:J4")
            ws["A4"].font = Font(italic=True, color="808080")
            ws["A4"].alignment = Alignment(horizontal="center")

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col_letter].width = 15
